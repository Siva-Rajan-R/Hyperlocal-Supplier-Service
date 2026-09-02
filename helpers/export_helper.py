import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Any

def generate_csv_bytes(headers: List[str], rows: List[List[Any]]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([str(v) if v is not None else "" for v in row])
    return output.getvalue().encode('utf-8-sig')

def generate_xlsx_bytes(headers: List[str], rows: List[List[Any]], sheet_name: str = "Export") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row in rows:
        ws.append([v if v is not None else "" for v in row])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
